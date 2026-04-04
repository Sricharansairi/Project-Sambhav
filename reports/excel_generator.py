import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import os

def generate_xlsx(data, output_path):
    """
    Professional XLSX generator for Project Sambhav.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Prediction Report"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    center_align = Alignment(horizontal="center")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # 1. Summary
    ws.append(["PROJECT SAMBHAV - PREDICTION REPORT"])
    ws.merge_cells('A1:B1')
    ws['A1'].font = Font(size=14, bold=True)
    ws.append([])
    
    summary = [
        ["Prediction ID", data.get("prediction_id", "N/A")],
        ["Domain", data.get("domain", "N/A")],
        ["Generated At", data.get("generated_at", "N/A")],
        ["Mode", data.get("mode", "guided").title()]
    ]
    for row in summary:
        ws.append(row)
    ws.append([])

    # 2. Probabilities
    ws.append(["PROBABILISTIC INFERENCE LAYERS"])
    ws.merge_cells(f'A{ws.max_row}:D{ws.max_row}')
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    
    headers = ["Layer", "Probability", "Contribution", "Status"]
    ws.append(headers)
    for cell in ws[ws.max_row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    layers = [
        ["ML Prediction Layer", data.get("ml_probability", "N/A"), "65%", "CALIBRATED"],
        ["LLM Inference Layer", data.get("llm_probability", "N/A"), "35%", "STOCHASTIC"],
        ["Reconciled Final", data.get("reconciled_probability", "N/A"), "100%", "OPTIMIZED"]
    ]
    for layer in layers:
        ws.append(layer)
    ws.append([])

    # 3. Outcomes
    ws.append(["DETAILED OUTCOMES"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.append(["Outcome Label", "Probability Score"])
    for cell in ws[ws.max_row]:
        cell.font = header_font
        cell.fill = header_fill
    
    outcomes = data.get("outcomes", [])
    for o in outcomes:
        ws.append([o.get("label", ""), o.get("probability", "")])
    ws.append([])

    # 4. Parameters
    ws.append(["INPUT PARAMETERS"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.append(["Parameter", "Value"])
    for cell in ws[ws.max_row]:
        cell.font = header_font
        cell.fill = header_fill
        
    params = data.get("parameters", {})
    for k, v in params.items():
        ws.append([k.replace("_", " ").title(), str(v)])
    ws.append([])

    # 5. SHAP
    ws.append(["FEATURE CONTRIBUTION (SHAP)"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.append(["Feature", "Impact Score", "Direction"])
    for cell in ws[ws.max_row]:
        cell.font = header_font
        cell.fill = header_fill
        
    shaps = data.get("shap_values", {})
    for k, v in shaps.items():
        direction = "Positive (+)" if v > 0 else "Negative (-)"
        ws.append([k.replace("_", " ").title(), v, direction])
    ws.append([])

    # Column Width Adjustment
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 5

    wb.save(output_path)
